from product.form import CategoryForm, AddProductForm, AddSalesForm, AddPurchaseForm
from datetime import datetime, date
import time
from authentication.models import CustomUser
from django.shortcuts import render, redirect, get_object_or_404
from .utils import generate_random_code
from product.models import *
from django.http import JsonResponse, HttpResponse
from .form import FileUploadForm
import pandas as pd
from django.views.decorators.csrf import csrf_exempt
import json
from django.db.models import DateTimeField
from django.db.models import (
    ExpressionWrapper,
    F,
    Func,
    Value,
    Sum,
    DateField,
    CharField,
    FloatField,
    DecimalField,
)
from django.db.models.functions import Trunc, Cast, TruncDate
from django.db.models import Q


from django.template.loader import render_to_string
from django.views import generic
from xhtml2pdf import pisa
from io import BytesIO
import re

import xlwt

# ???  decorator
from django.contrib.auth.decorators import login_required, user_passes_test

# openpyxl
from openpyxl import load_workbook
import uuid

# from barcode import Code128, Code39
# from barcode.writer import ImageWriter
import uuid
from django.conf import settings
import os
from django.db.models.functions import TruncMonth
from collections import defaultdict


def is_admin(user):
    print(user)
    return user.groups.filter(name="ADMIN").exists()


@login_required
def dashboard(request):
    # get the value of all your store products at cost price
    total_cost = Products.objects.aggregate(
        total_cost=Sum(F("unit_price") * F("quantity_in_stock"))
    )["total_cost"]
    # formatted_total_cost = '{:,.0f}'.format(total_cost)
    total_cost_at_selling_price = Products.objects.aggregate(
        total_cost=Sum(F("cost") * F("quantity_in_stock"))
    )["total_cost"]
    total_cost_of_sales = Sales.objects.aggregate(total_cost=Sum(F("total")))[
        "total_cost"
    ]

    total_stock_cost = (
        Products.objects.filter(stock_take_done=True)
        .annotate(new__stock=Cast("new_stock", FloatField()))
        .aggregate(purchase_cost=Sum(F("unit_price") * F(("quantity_in_stock"))))
    )
    print(total_stock_cost)
    total_no_of_products = Products.objects.all().count()
    total_no_of_sales = Sales.objects.all().count()
    # sales_of_yesterday = Sales.objects.filter(
    #     date_sold__date=datetime.date.today() - datetime.timedelta(days=1)
    # ).aggregate(total_sales=Sum(F("total")))["total_sales"]
    sales_of_yesterday = Sales.objects.filter(
        date_sold__exact=datetime.date.today() - datetime.timedelta(days=1)
    ).aggregate(total_sales=Sum(F("total")))["total_sales"]

    # todays sales

    sales_of_today = Sales.objects.filter(
        date_sold__exact=datetime.date.today()
    ).aggregate(total_sales=Sum(F("total")))["total_sales"]

    # ? top selling products
    top_selling_products = Products.objects.annotate(
        total_quantity_sold=Sum("sales__quantity")
    ).order_by("-total_quantity_sold")[:5]
    print(top_selling_products)

    # value_if_true if condition else value_if_false
    total_sales_of_yesterday = sales_of_yesterday if sales_of_yesterday else 0
    total_sales_of_today = sales_of_today if sales_of_today else 0
    if total_cost_of_sales is not None:
        formatted_total_cost_of_sales = format(int(total_cost_of_sales), ",d")
    else:
        formatted_total_cost_of_sales = 0

    if total_cost is not None:
        formatted_total_cost = format(int(total_cost), ",d")
    else:
        formatted_total_cost = 0

    if total_cost_at_selling_price is not None:
        foramtted_total_cost_at_selling_price = format(
            int(total_cost_at_selling_price), ",d"
        )
    else:
        foramtted_total_cost_at_selling_price = 0

    context = {
        "total_cost": formatted_total_cost,
        "total_cost_at_selling_price": foramtted_total_cost_at_selling_price,
        "total_sales_of_yesterday": format(int(total_sales_of_yesterday), ",d"),
        "total_sales_of_today": format(int(total_sales_of_today), ",d"),
        "total_no_of_products": total_no_of_products,
        "total_no_of_sales": total_no_of_sales,
        "total_cost_of_sales": formatted_total_cost_of_sales,
        "top_selling_products": top_selling_products,
        "total_stock_cost": format(int(total_stock_cost["purchase_cost"]), ",d"),
    }
    return render(request, "dashboard/dashboard.html", context)


# ?------------------------------------ADD PRODUCT ----------------------------
class AddProduct(generic.TemplateView):
    template_name = "products/add_products.html"


class ProductList(generic.View):
    template_name = "products/products_list.html"
    model = Products
    context_object_name = "products"
    ordering = ["created_at"]

    #  product form
    def get(self, request, *args, **kwargs):
        products = self.model.objects.all()
        products_form = AddProductForm()
        context = {"products": products, "products_form": products_form}
        return render(request, self.template_name, context)

    # def post(self, request, *args, **kwargs ):


def get_single_product(request, pk):
    # if request.method == "GET"
    prod_id = str(pk)
    print(prod_id)

    # Assuming 'id' is the field in your Products model that you want to filter on
    product = get_object_or_404(Products, id=prod_id)
    product_data = {
        "product_name": product.product_name,
        "unit_price": product.unit_price,
        "cost": product.cost,
        "quantity_in_stock": product.quantity_in_stock,
        "new_stock": product.new_stock,
    }

    print(product_data)
    return JsonResponse(product_data)


@login_required
def addproduct(request):
    form = AddProductForm()
    uploadform = FileUploadForm()
    context = {"form": form, "uploadform": uploadform}

    return render(request, "products/add_products.html", context)


@login_required
def list_category(request):
    categories = Category.objects.all()
    context = {"categories": categories}

    return render(request, "categories/list_cat.html", context)


@login_required
def AddCategory(request):
    error = None
    if request.method == "POST":
        form = CategoryForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
        else:
            error = form.errors
    else:
        form = CategoryForm()
    context = {"form": form, "error": error}
    return render(request, "categories/add_cat.html", context)


def add_single_product(request):
    if request.method == "POST":
        print(request.POST, request.FILES)
        product_name = request.POST["product_name"]
        category = request.POST["category_id"]
        cost_price = request.POST["unit_price"]
        selling_price = request.POST["cost"]
        quantity_in_stock = request.POST["quantity_in_stock"]
        brand_name = request.POST["brand"]
        unit_of_measure = request.POST.get("unit_of_measure", "pcs")
        reorder_level = request.POST.get("reorder_level", 0)
        description = request.POST.get("description", "none")
        # product_image = request.FILES["pic"]

        category = Category.objects.get(id=category)
        brand = Brand.objects.get(id=brand_name)
        # gererate product code

        product_code = generate_random_code(13)

        Products.objects.create(
            product_name=product_name,
            category_id=category,
            cost=selling_price,
            unit_price=cost_price,
            quantity_in_stock=quantity_in_stock,
            brand=brand,
            unit_of_measure=unit_of_measure,
            reorder_level=reorder_level,
            description=description,
            product_code=product_code,
            stock_take_done=True,
            # product_image=product_image,
        )
        return redirect("products:product_list")

    return render(request, "products/add_products.html")


# ? ------------------------------------- sales -------------------------------------


def add_single_sell(request):
    form = AddSalesForm
    if request.method == "POST":
        user = request.user
        data = request.POST
        data_sold = data["date_sold"]
        product_name = data["product_name"]
        product_uuid = data["product_uuid"]
        discount = data["discount"]
        quantity = int(data["quantity"])
        price = int(data["price"])
        product = Products.objects.get(product_code=product_uuid)
        product.stock_take_done = True
        total_price = (int(price) * int(quantity)) - int(discount)
        print(total_price, discount)
        Sales.objects.create(
            user=user,
            date_sold=data_sold,
            product=product,
            quantity=quantity,
            total=total_price,
            price=price,
            discount=discount,
        )
        return redirect("products:add_single_sell")

    return render(
        request,
        "sales/add_single_sell.html",
        {"form": form},
    )


class AddSale(generic.View):
    template_name = "sales/add_sales.html"

    def get(self, request, *args, **kwargs):
        return render(request, self.template_name)

    def post(self, request, *args, **kwargs):
        #! if request.is_ajax and request.method == "POST":
        #! checkin if request is ajax
        # if request.META.get("HTTP_X_REQUESTED_WITH") == "XMLHttpRequest":
        name = request.POST.get("product")
        res = None

        # qs = Products.objects.filter(product_name__icontains=name or product_code__icontains=name)
        qs = Products.objects.filter(
            Q(product_name__icontains=name) | Q(product_code__icontains=name)
        )

        data = []
        print(qs)
        if len(qs) > 0 and len(name) > 0:
            for i in qs:
                item = {
                    "code": i.product_code,
                    "name": i.product_name,
                    "price": i.cost,
                    "description": i.description,
                    "purchase_price": i.unit_price,
                }
                data.append(item)
            res = data
        else:
            res = "No games found ........"
            return JsonResponse({"data": "Not found"}, status=400)
        return JsonResponse({"data": res}, status=200)


def complete_sale(request):
    items_sold = []
    if request.method == "POST":
        print(request.POST["cart"])
        cart = request.POST["cart"]
        # convert to json
        cart_data = json.loads(cart)

        # convert to dictionary
        # cart = eval(cart)
        # loop through a dictionary
        n = 0

        for key, value in cart_data.items():
            n += 1
            print(n)
            quantity = int(value["quantity"])
            product_uuid = value["product_uuid"]
            total_price = (int(value["sales_price"])) * int(quantity) - int(
                value["discount"]
            )
            discount = value["discount"]
            price = int(value["sales_price"])
            user = CustomUser.objects.get(username=request.user.username)
            print(quantity)
            # get product by uuid and quantity
            product = Products.objects.get(product_code=product_uuid)
            # check if product quantity is less than quantity
            if int(product.quantity_in_stock) < int(quantity):
                print("Not enough stock")
                message = "Not enough stock"

                return JsonResponse({"data": "Not enough stock"}, status=400)
            print(user, product)
            # create a sale
            Sales.objects.create(
                user=user,
                product=product,
                quantity=quantity,
                total=total_price,
                discount=discount,
                price=price,
            )
            # create the receipt data
            items_sold.append(
                {
                    "product_name": product,
                    "quantity": quantity,
                    "total_price": total_price,
                    "price": price,
                    "cashiers_name": user.full_name,
                },
            )
    elif request.method == "GET":
        items_dict = [item for item in items_sold]
        return render(request, "sales/receipt.html", {"items": items_dict})

    return JsonResponse({"data": "success"}, status=200)

    # Generate the PDF file
    # response = HttpResponse(content_type="application/pdf")
    # response["Content-Disposition"] = 'filename="receipt.pdf"'

    # # Generate the PDF content using the HTML template

    # pdf = BytesIO()
    # pisa.CreatePDF(BytesIO(html.encode("UTF-8")), pdf)
    # response.write(pdf.getvalue())

    # return response

    #         redirect("products:list_sales")

    # return JsonResponse({"data": "success"}, status=200)


class ListSale(generic.View):
    def get(self, request, *args, **kwargs):
        sales = (
            Sales.objects.all()
            .order_by("-date_sold")
            .annotate(
                total_profit=Sum((F("price") - F("current_cost_price")) * F("quantity"))
            )  # .annotate(actual_stock_left=F("product__quantity_in_stock"))
        )
        # edit date format to Tuesday 09 March 2021
        # my_sales = []
        # for sale in sales:
        #     sale.date_sold = sale.date_sold.strftime("%A %d %B %Y")
        #     print(sale.date_sold)
        #     my_sales.append(sale)

        # # order my_sales list by date_sold such that thursday 11 march 2021 come after wednesday 10 march 2021
        # my_sales = sorted(
        #     my_sales,
        #     key=lambda x: x.date_sold.date(),
        #     reverse=True,
        # )

        # print(my_sales)

        todays_total_sales = Sales.objects.filter(date_sold=datetime.date.today())
        todays_total_sales = sum([sale.total for sale in todays_total_sales])
        todays_total_sales = "{:,.0f}".format(todays_total_sales)

        context = {"sales": sales, "todays_total_sales": todays_total_sales}
        return render(request, "sales/list_sales.html", context)


# ? --------------------------- Purchases ---------------------------
class AddPurchase(generic.View):
    form_class = FileUploadForm
    form_purchase = AddPurchaseForm
    products_form = AddProductForm()

    def get(self, request, *args, **kwargs):
        form = self.form_class()
        form_purchase = self.form_purchase()
        return render(
            request,
            "purchases/add_purchases.html",
            {
                "form": form,
                "form_purchase": form_purchase,
                "products_form": self.products_form,
            },
        )


def add_single_purchase(request):
    form = FileUploadForm
    form_purchase = AddPurchaseForm
    products_form = AddProductForm()
    if request.method == "POST":
        product = Products.objects.get(product_code=request.POST.get("product_uuid"))
        quantity = request.POST.get("quantity")
        purchase_price = request.POST.get("purchase_price")
        supplier = Supplier.objects.get(id=request.POST.get("supplier"))
        purchase_date = request.POST.get("purchase_date")
        purchase = Purchases.objects.create(
            product=product,
            quantity=quantity,
            purchase_price=purchase_price,
            purchase_date=purchase_date,
            supplier=supplier,
        )
        purchase.save()
        product = Products.objects.get(product_code=request.POST.get("product_uuid"))
        product.unit_price = purchase_price
        product.quantity_in_stock += int(quantity)
        product.save()

        return redirect("products:add_purchases")
    print(form_purchase)
    return render(
        request,
        "purchases/add_purchases.html",
        {"form": form, "form_purchase": form_purchase, "products_form": products_form},
    )


class ListPurchase(generic.ListView):
    model = Purchases
    queryset = Purchases.objects.all()
    template_name = "purchases/list_purchases.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["purchases"] = Purchases.objects.all().annotate(
            total_quantity=F("quantity") * F("purchase_price")
        )
        print(context["purchases"])

        return context


def generate_purchase_list(request):
    product = Products.objects.all().order_by("reorder_level")

    context = {"products": product}
    return render(request, "purchases/generate_purchase_list.html", context)


def add_purchase_upload(request):
    if request.method == "POST":
        data = request.POST
        # get values form the dict

        product_name = request.POST.getlist("product_name")
        product_code = request.POST.getlist("product_code")
        description = request.POST.getlist("description")
        quantity = request.POST.getlist("quantity")
        purchase_price = request.POST.getlist("purchase_price")
        sales_price = request.POST.getlist("sales_price")
        purchase_quantity = request.POST.getlist("purchase_quantity")
        purchase_date = request.POST.getlist("purchase_date")
        supplier = request.POST.getlist("supplier")
        # print(product_name,product_code,description,quantity,purchase_date,purchase_price,purchase_quantity,supplier,sales_price)
        for (
            product_name,
            product_code,
            description,
            quantity,
            purchase_date,
            purchase_price,
            purchase_quantity,
            supplier,
            sales_price,
        ) in zip(
            product_name,
            product_code,
            description,
            quantity,
            purchase_date,
            purchase_price,
            purchase_quantity,
            supplier,
            sales_price,
        ):
            # check if product exists
            prod = Products.objects.filter(product_code=product_code)
            if prod.exists():
                print("{} exists".format(product_name))
                new_quantity = int(quantity) + int(purchase_quantity)
                my_product = Products.objects.get(product_code=product_code)
                # get current quantity and update it with new quantity
                my_product.quantity_in_stock = int(my_product.quantity_in_stock) + int(
                    new_quantity
                )
                my_product.new_arrival = True
                my_product.save()

                if not Supplier.objects.filter(name=supplier).exists():
                    Supplier.objects.create(name=supplier)
                date = time.strptime(
                    purchase_date.replace(", midnight", ""), "%B %d, %Y"
                )
                date = time.strftime("%Y-%m-%dT%H:%M:%S", date)

                Purchases.objects.create(
                    product=my_product,
                    quantity=purchase_quantity,
                    purchase_price=purchase_price,
                    supplier=Supplier.objects.get(name=supplier),
                    purchase_date=date,
                )

            elif not prod.exists():
                new_quantity = int(quantity) + int(purchase_quantity)
                print(new_quantity)
                new_product = Products.objects.create(
                    product_name=product_name,
                    product_code=product_code,
                    description=description,
                    unit_of_measure="pcs",
                    category_id=Category.objects.get(category_name="others"),
                    quantity_in_stock=new_quantity,
                    unit_price=purchase_price,
                    cost=sales_price,
                    reorder_level=0,
                    brand=Brand.objects.get(brand_name="others"),
                    new_arrival=True,
                )
                print("{} created and doesnot exits".format(product_name))
                new_product.save()
                if not Supplier.objects.filter(name=supplier).exists():
                    Supplier.objects.create(name=supplier)
                date = time.strptime(
                    purchase_date.replace(", midnight", ""), "%B %d, %Y"
                )
                date = time.strftime("%Y-%m-%dT%H:%M:%S", date)
                Purchases.objects.create(
                    product=new_product,
                    quantity=purchase_quantity,
                    purchase_price=purchase_price,
                    purchase_date=date,
                    supplier=Supplier.objects.get(name=supplier),
                )
                # return redirect("products:list_products")

    return redirect("products:product_list")


def upload_purchase(request):
    form = FileUploadForm()
    context = {"form": form}
    if request.method == "POST":
        # print(request.POST)
        file = request.FILES["file"]

        # sheet
        wb = load_workbook(file)
        # get all sheet names
        sheets = wb.get_sheet_names()
        context = {"products": []}

        for sheet in sheets:
            # get sheet by name
            ws = wb.get_sheet_by_name(sheet)
            values = [row[0:] for row in ws.values]
            # convert to dataframe

            df = pd.DataFrame(values)

            header_list = [
                "Product Code",
                "Product Name",
                "Description",
                "Quantity",
                "Cost Price",
                "Sales Price",
                "Reorder Level",
                "Purchase Quantity",
                "Purchase Date",
                "Supplier",
            ]
            # print(df.iloc[0].tolist())
            if df.iloc[0].tolist() == header_list:
                df = df[0:]
                # get the first row as header
                df.columns = df.iloc[0]
                # drop the first row
                df = df[1:]
                # reset index
                df = df.reset_index(drop=True)
                # drop nonwe in PURCHASE PRICE, SALES PRICE, QUANTITY, PRODUCT NAME
                df = df.dropna(
                    subset=[
                        "Cost Price",
                        "Sales Price",
                        "Purchase Quantity",
                        "Purchase Date",
                        "Supplier",
                    ],
                    how="all",
                )

                # convert to dictionary
                product_name = df["Product Name"].tolist()
                product_code = df["Product Code"].tolist()
                description = df["Description"].tolist()
                quantity = df["Quantity"].tolist()
                cost_price = df["Cost Price"].tolist()
                sales_price = df["Sales Price"].tolist()
                reorder_level = df["Reorder Level"].tolist()
                purchase_quantity = df["Purchase Quantity"].tolist()
                purchase_date = df["Purchase Date"].tolist()
                supplier = df["Supplier"].tolist()

                for i in range(len(product_name)):
                    # check if product exists
                    product = Products.objects.filter(product_code=product_code[i])
                    if product.exists():
                        product = product[0]
                        products = {
                            "product_name": product_name[i],
                            "product_code": product_code[i],
                            "description": description[i],
                            "quantity": quantity[i],
                            "cost_price": cost_price[i],
                            "sales_price": sales_price[i],
                            "reorder_level": reorder_level[i],
                            "purchase_quantity": purchase_quantity[i],
                            "purchase_date": purchase_date[i],
                            "supplier": supplier[i],
                            "new_quantity": product.quantity_in_stock
                            + int(purchase_quantity[i]),
                            "new_product": "False",
                        }
                        print(products)
                        context["products"].append(products)
                        print(context)
                    else:
                        products = {
                            "product_code": generate_random_code(13),
                            "product_name": product_name[i],
                            "description": description[i],
                            "quantity": quantity[i],
                            "cost_price": cost_price[i],
                            "sales_price": sales_price[i],
                            "reorder_level": reorder_level[i],
                            "purchase_quantity": purchase_quantity[i],
                            "purchase_date": purchase_date[i],
                            "supplier": supplier[i],
                            "new_product": "True",
                        }
                        print(products)
                        context["products"].append(products)
                return render(request, "purchases/add_purchases.html", context)

            # else:
            #     products = {
            #         "product_name": product_name[i],
            #         "product_code": product_code[i],
            #         "description": description[i],
            #         "quantity": quantity[i],
            #         "cost_price": cost_price[i],
            #         "sales_price": sales_price[i],
            #         "reorder_level": reorder_level[i],
            #         "purchase_quantity": purchase_quantity[i],
            #         "purchase_date": purchase_date[i],
            #         "supplier": supplier[i],
            #         "new_product": "True",
            #     }

            #     context["products"].append(products)
            #     print(context)

    return render(request, "purchases/add_purchases.html", context)


@login_required
def upload_products(request):
    if request.method == "POST":
        form = FileUploadForm(request.POST, request.FILES)
        if form.is_valid():
            file = form.cleaned_data["file"]
            # sheet
            wb = load_workbook(file)
            # get all sheet names
            sheets = wb.get_sheet_names()
            context = {"products": []}
            for sheet in sheets:
                # get sheet by name
                ws = wb.get_sheet_by_name(sheet)
                values = [row[1:] for row in ws.values]
                # convert to dataframe

                df = pd.DataFrame(values)

                header_list = [
                    "PRODUCT NAME",
                    "DESCRIPTION",
                    "QUANTITY",
                    "PURCHASE PRICE",
                    "SALES PRICE",
                    "CATEGORY",
                    "UNIT ID",
                    "BRAND",
                ]
                # check if header is in the first row

                if df.iloc[0].tolist() == header_list:
                    df = df[0:]
                    # get the first row as header
                    df.columns = df.iloc[0]
                    # drop the first row
                    df = df[1:]
                    # reset index
                    df = df.reset_index(drop=True)
                    # drop nonwe in PURCHASE PRICE, SALES PRICE, QUANTITY, PRODUCT NAME
                    df = df.dropna(
                        subset=[
                            "PURCHASE PRICE",
                            "SALES PRICE",
                            "QUANTITY",
                            "PRODUCT NAME",
                            "CATEGORY",
                            "UNIT ID",
                        ]
                    )
                    # get the first row as header
                    product_name = df["PRODUCT NAME"].tolist()
                    description = df["DESCRIPTION"].tolist()
                    quantity = df["QUANTITY"].tolist()
                    purchase_price = df["PURCHASE PRICE"].tolist()
                    sales_price = df["SALES PRICE"].tolist()
                    category = df["CATEGORY"].tolist()
                    unit = df["UNIT ID"].tolist()
                    brand = df["BRAND"].tolist()

                    for i in range(len(product_name)):
                        # create uuid
                        code = generate_random_code(13)
                        print(code)

                        # create context

                        products = {
                            "uuid": code,
                            "product_name": product_name[i],
                            "description": description[i],
                            "quantity": quantity[i],
                            "purchase_price": purchase_price[i],
                            "sales_price": sales_price[i],
                            "category": category[i],
                            "unit": unit[i],
                            "brand": brand[i],
                        }

                        context["products"].append(products)
                    return render(request, "products/add_products.html", context)

                else:
                    print("header not found")
    return render(request, "products/add_products.html")


def generate_barcode(request):
    product = Products.objects.all()
    for uuid in product:
        code = uuid.product_code
        product_name = uuid.product_name.replace(" ", "_").replace("*", "_")
        barcode = Code128(code, writer=ImageWriter())

        # Set the width of the barcode image to 600 pixels
        barcode.writer.set_options({"width": 600})

        # Save the barcode image to a file

        # save the barcode to a file in the images directory
        image_dir = os.path.join(settings.BASE_DIR, "images")
        if not os.path.exists(image_dir):
            os.makedirs(image_dir)
        # change image name to product name
        barcode.save(
            os.path.join(image_dir, product_name),
            options={"background": "white", "foreground": "black"},
        )
    return JsonResponse({"success": True})


def add_product_upload(request):
    if request.method == "POST":
        data = request.POST
        # get values form the dict
        for i in request.POST.getlist("product_name"):
            product_name = request.POST.getlist("product_name")
            description = request.POST.getlist("description")
            quantity = request.POST.getlist("quantity")
            purchase_price = request.POST.getlist("purchase_price")
            sales_price = request.POST.getlist("sales_price")
            category = request.POST.getlist("category")
            unit = request.POST.getlist("unit")
            brand = request.POST.getlist("brand")
            uuid = request.POST.getlist("uuid")
            # print(product_name, description, quantity, purchase_price, sales_price, category, unit, brand, uuid)
        for (
            product_name,
            description,
            quantity,
            purchase_price,
            sales_price,
            category,
            unit,
            brand,
            uuid,
        ) in zip(
            product_name,
            description,
            quantity,
            purchase_price,
            sales_price,
            category,
            unit,
            brand,
            uuid,
        ):
            print(
                product_name,
                description,
                quantity,
                purchase_price,
                sales_price,
                category,
                unit,
                brand,
                uuid,
            )
            # delte all prod
            # Products.objects.all().delete()
            # Category.objects.all().delete()
            # Brand.objects.all().delete()
            if Category.objects.filter(category_name=category).exists() == False:
                category = Category.objects.create(category_name=category)
                category.save()
            else:
                category = Category.objects.get(category_name=category)

            if Brand.objects.filter(brand_name=brand).exists() == False:
                brand = Brand.objects.create(brand_name=brand)
                brand.save()
            else:
                brand = Brand.objects.get(brand_name=brand)
                brand.save()

            products = Products.objects.create(
                product_name=product_name,
                description=description,
                quantity_in_stock=quantity,
                unit_price=purchase_price,
                cost=sales_price,
                category_id=category,
                unit_of_measure="pcs",
                brand=brand,
                product_code=uuid,
            )
            products.save()

        return JsonResponse({"success": True})
    return JsonResponse({"success": False})


def get_product_by_uuid(request, uuid):
    if request.method == "GET":
        product = Products.objects.get(product_code=uuid)

        data = {
            "product_name": product.product_name,
            "sales_price": product.cost,
            "product_uuid": product.product_code,
            "purchase_price": product.unit_price,
        }
        print(data)
        return JsonResponse(data)


@login_required
def products_list(request):
    products = Products.objects.all()
    context = {"products": products}
    return render(request, "products/products_list.html", context)


# ?------------------REPORTS----------------------------------------


def reports(request):
    # ? getting a single date  form the purchases query set
    purchases_dates = (
        Purchases.objects.all()
        .values("purchase_date")
        .order_by("purchase_date")
        .distinct()
    )

    # ? weekly profit
    from django.db.models.functions import TruncWeek

    weekly_profit = (
        Sales.objects.annotate(week_start=TruncWeek("timestamp"))
        .values("week_start")
        .annotate(
            total_profit=Sum((F("price") - F("current_cost_price")) * F("quantity"))
            # total_profit=Sum(
            # (F("price") - F("current_cost_price")) * F("quantity") - F("discount") * F("quantity")
            # )
        )
        .order_by("week_start")
    )

    for week_profit in weekly_profit:
        week_profit["week_start"] = week_profit["week_start"].strftime("%A, %B %d, %Y")
        # week_profit["week_start_orgi"] = week_profit["week_start"].strftime(
        #     "%A, %B %d, %Y"
        # )
    print(weekly_profit)
    if purchases_dates:
        for date in purchases_dates:
            # multiping the quantity and purchase_price to get the total
            # print(date["purchase_date"])
            # purchases_sum = Purchases.objects.filter(
            #     purchase_date=date["purchase_date"]
            # ).aggregate(total_purchase_price=Sum(F("quantity") * F("purchase_price")))
            # context = {
            #     "purchases": {
            #         "purchases_dates": date["purchase_date"],
            #         "purchases_sum": format(purchases_sum["total_purchase_price"], ",d"),
            #     }
            # }
            purchases = (
                Purchases.objects.values("purchase_date")
                .annotate(total_purchase_price=Sum(F("quantity") * F("purchase_price")))
                .order_by("purchase_date")
            )
            for purchase in purchases:
                # formating date to "Tuesday, April 25, 2023"
                purchase["purchase_date"] = purchase["purchase_date"].strftime(
                    "%A, %B %d, %Y"
                )

    # purchases_sum = Purchases.objects.aggregate(Sum("total"))
    return render(
        request,
        "reports/reports.html",
        {"purchases": purchases, "weekly_profit": weekly_profit},
    )


def daily_profit(request):
    # Sales.objects.annotate(week_start=TruncWeek("timestamp"))
    #     .values("week_start")
    #     .annotate(
    #         total_profit=Sum((F("price") - F("current_cost_price")) * F("quantity"))
    #         # total_profit=Sum(
    #         # (F("price") - F("current_cost_price")) * F("quantity") - F("discount") * F("quantity")
    #         # )
    #     )
    #     .order_by("week_start")

    # daily_profit = (
    #     Sales.objects.annotate(day=TruncDate("timestamp"))
    #     .values("day")
    #     .annotate(
    #         total_profit=Sum(
    #             ExpressionWrapper(
    #                 (F("price") - F("current_cost_price")) * F("quantity"),
    #                 output_field=DecimalField(),
    #             )
    #         )
    #     )
    #     .order_by("day")
    # )
    # total_profit_per_day = {}
    # for item in daily_profit:
    #     total_profit_per_day[item["day"]] = item["total_profit"]

    # print(total_profit_per_day)
    # Calculate the daily profit and list all products sold with their individual profits
    daily_sales = (
        Sales.objects.annotate(day=TruncDate("timestamp"))
        .values("day", "product")
        .annotate(
            profit=ExpressionWrapper(
                (F("price") - F("current_cost_price")) * F("quantity"),
                output_field=DecimalField(),
            )
        )
        .order_by("day", "product")
        .annotate(product_name=F("product__product_name"))
    )

    # Calculate the total profit for each day
    total_profit_per_day = {}
    for item in daily_sales:
        item['day'] = item['day'].strftime("%Y/%m/%d")
        day = item["day"]
        profit = item["profit"]
        total_profit_per_day[day] = total_profit_per_day.get(day, 0), + profit,
        # total_profit_per_day = total_profit_per_day.get(day, 0), + profit,

    # Print the daily sales with individual profits and total profit for each day
    for item in daily_sales:
        print(
            f"Day: {item['day']} | Product: {item['product_name']} | Profit: {item['profit']} | Total:{total_profit_per_day}"
        )
        # If you want to save this data, you can store it in a list or dictionary

    print("\nTotal Profit per Day:")
    for day, total_profit in total_profit_per_day.items():
        print(f"Day: {day} | Total Profit: {total_profit}")
    current_month = datetime.datetime.now().replace(day=1)
    total_profit_for_month = Sales.objects.filter(
        timestamp__month=current_month.month
    ).aggregate(
        total_profit=Sum(
            ExpressionWrapper(
                (F("price") - F("current_cost_price")) * F("quantity"),
                output_field=DecimalField(),
            )
        )
    )[
        "total_profit"
    ]

    context = {
        "daily_sales": daily_sales,
        "total_profit_per_day": total_profit_per_day,
        "total_profit_for_month": total_profit_for_month,
    }
    print(context)
    return render(request, "reports/daily_profit.html", context)


def get_profit_for_single_date(request, pk):

    pass



def get_month_profit(request):

    monthly_profit = defaultdict(dict)

    # Assuming Sales model has 'price', 'current_cost_price', 'quantity', 'product', and 'timestamp' fields
    # daily_sales = (
    #     Sales.objects.annotate(month=TruncMonth("timestamp"))
    #     .values("month", "product")
    #     .annotate(
    #         profit=ExpressionWrapper(
    #             (F("price") - F("current_cost_price")) * F("quantity"),
    #             output_field=DecimalField(),
    #         )
    #     )
    #     .order_by("month", "product")
    #     .annotate(product_name=F("product__product_name"))
    # )

    # for sale in daily_sales:
    #     month = sale["month"].strftime("%Y-%m")
    #     if month not in monthly_profit:
    #         monthly_profit[month] = {}
    #     if sale["product_name"] not in monthly_profit[month]:
    #         monthly_profit[month][sale["product_name"]] = 0
    #     monthly_profit[month][sale["product_name"]] += sale["profit"]

    # context = {

    # 'monthly_profit': monthly_profit
    # }
    # print(dict(monthly_profit))
    monthly_profits = (
    Sales.objects.annotate(month=TruncMonth("timestamp"))
            .values("month")
            .annotate(
            total_profit=Sum(
                ExpressionWrapper(
                    (F("price") - F("current_cost_price")) * F("quantity"),
                    output_field=DecimalField(),
                )
            )
    )
    .order_by("month")
    )
    # return render(request, 'reports/month_profit.html', context)


    return render(request, 'reports/month_profit.html', {'monthly_profits': monthly_profits})


    

# ?  ----------export to excel----------------


def export_to_excel(request):
    response = HttpResponse(content_type="application/ms-excel")
    response["Content-Disposition"] = 'attachment; filename="products.xls"'
    wb = xlwt.Workbook(encoding="utf-8")
    ws = wb.add_sheet("Products")
    row_num = 0
    font_style = xlwt.XFStyle()
    font_style.font.bold = True
    columns = [
        "Product Code",
        "Product Name",
        "Description",
        "Quantity",
        "Cost Price",
        "Sales Price",
        "Reorder Level",
        "Purchase Quantity",
        "Purchase Date",
        "Supplier",
    ]

    for col_num in range(len(columns)):
        ws.write(row_num, col_num, columns[col_num], font_style)

    font_style = xlwt.XFStyle()
    rows = Products.objects.all().values_list(
        "product_code",
        "product_name",
        "description",
        "quantity_in_stock",
        "unit_price",
        "cost",
        "reorder_level",
    )

    for row in rows:
        row_num += 1
        for col_num in range(len(row)):
            ws.write(row_num, col_num, row[col_num], font_style)

    wb.save(response)
    return response


def export_all_products_excel(request):
    print(request)

    response = HttpResponse(content_type="application/ms-excel")
    response["Content-Disposition"] = 'attachment; filename="products.xls"'
    wb = xlwt.Workbook(encoding="utf-8")
    ws = wb.add_sheet("AllProducts")
    row_num = 0
    font_style = xlwt.XFStyle()
    font_style.font.bold = True
    columns = [
        "Product Code",
        "Product Name",
        "Description",
        "Unit",
        "Quantity",
        "Purchase Price",
        "Sales Price",
        "Reorder Level",
        "Category",
        "Brand",
        "Image",
    ]
    for col_num in range(len(columns)):
        ws.write(row_num, col_num, columns[col_num], font_style)

    font_style = xlwt.XFStyle()
    rows = (
        Products.objects.all()
        .values_list(
            "product_code",
            "product_name",
            "description",
            "unit_of_measure",
            "quantity_in_stock",
            "unit_price",
            "cost",
            "reorder_level",
        )
        .annotate(
            category_name=F("category_id__category_name"),
            brand_name=F("brand__brand_name"),
            product_image=F("product_image"),
        )
        .order_by("product_name")
    )
    for row in rows:
        row_num += 1
        for col_num in range(len(row)):
            ws.write(row_num, col_num, row[col_num], font_style)

    wb.save(response)
    return response


def export_all_sales_excel(request):
    response = HttpResponse(content_type="application/ms-excel")
    response["Content-Disposition"] = 'attachment; filename="products.xls"'
    wb = xlwt.Workbook(encoding="utf-8")
    ws = wb.add_sheet("AllProducts")
    row_num = 0
    font_style = xlwt.XFStyle()
    font_style.font.bold = True
    columns = [
        "Price",
        "Discount",
        "Stock Left",
        "quantity",
        "Current Cost Price",
        "Product Code",
        "Product Name",
        "Date",
    ]
    for col_num in range(len(columns)):
        ws.write(row_num, col_num, columns[col_num], font_style)

    font_style = xlwt.XFStyle()
    list = (
        Sales.objects.all()
        .annotate(
            product_code=F("product__product_code"),
            product_name=F("product__product_name"),
        )
        .values_list(
            "date_sold",
            "price",
            "discount",
            "stock_left",
            "quantity",
            "current_cost_price",
            "product_code",
            "product_name",
        )
    )
    # getting date from the tuple and formating it to "Tuesday, April 25, 2023"
    rows = []
    for item in list:
        date = item[0].strftime("%A, %B %d, %Y")
        rows.append(item[1:] + (date,))

    for row in rows:
        row_num += 1
        for col_num in range(len(row)):
            ws.write(row_num, col_num, row[col_num], font_style)

    wb.save(response)
    return response
